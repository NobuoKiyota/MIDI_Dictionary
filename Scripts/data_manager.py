import os
import pandas as pd
import platform
import glob

class DataManager:
    def __init__(self, db_path_ignored, root_dir):
        # db_path is now mostly ignored or used as base pattern?
        # We will use root_dir to find library_*.xlsx
        self.root_dir = root_dir
        
        # Determine Local DB Name
        hostname = platform.node()
        self.local_db_name = f"{hostname}.xlsx"
        # self.local_db_path will be in MIDI_Library/[Hostname]/[Hostname].xlsx
        # But we need to make sure the folder exists when saving.
        # self.root_dir is base dir... MIDI_Library is inside base_dir?
        # Check main.py: self.lib_path = os.path.join(base_dir, "MIDI_Library")
        # And DataManager init gets self.db_path (library.xlsx) and self.base_dir.
        # So "root_dir" here corresponds to base_dir of the app?
        # Let's check main.py call: DataManager(self.db_path, self.base_dir)
        # Yes. So MIDI_Library is at os.path.join(self.root_dir, "MIDI_Library")
        self.midi_lib_path = os.path.join(self.root_dir, "MIDI_Library")
        self.host_dir = os.path.join(self.midi_lib_path, hostname)
        if not os.path.exists(self.host_dir):
            os.makedirs(self.host_dir)
            
        self.local_db_path = os.path.join(self.host_dir, self.local_db_name)
        
        self.columns = [
            "FileName", "FilePath", "Category", "Instruments", 
            "TimeSignature", "BAR", "Chord",
            "Root", "Group", "Comment"
        ]
        
        self.df = self.load_db()

    def load_db(self):
        # We don't load everything into self.df for the runtime "Active DB" usually?
        # Requirement: "Integrate master... at startup". 
        # So maybe load_db should just load *local* DB for appending?
        # BUT main.py uses data_manager.df for displaying the list.
        # IF we want to display EVERYTHING, we should load MasterLibrary.xlsx if it exists (which is result of integration).
        # OR we just load everything fresh every time?
        # "統合するマスターをMasterLibraly.xlsxとする... 統合処理はアプリ起動時に自動で行う"
        # (Master is MasterLibraly.xlsx. Integration is done auto on startup)
        #
        # So flow is:
        # 1. Integration runs -> creates MasterLibraly.xlsx from all sub-excels.
        # 2. Main App loads MasterLibraly.xlsx to show list?
        # 3. New entries -> Saved to Local Excel (and added to memory DF? or memory DF is Master?)
        
        # If we add to Master DF, next time we save, we must be careful NOT to overwrite Master with Local data only,
        # OR not to save Master data into Local file.
        
        # Proposed logic:
        # self.df ALWAYS represents what is on screen.
        # On Startup: self.df = MasterLibrary.xlsx (after integration)
        # On Add: Append to self.df AND Append to Local Excel.
        
        master_path = os.path.join(self.root_dir, "MasterLibraly.xlsx")
        if os.path.exists(master_path):
            try:
                df = pd.read_excel(master_path)
                
                # Consolidation: Merge 'Bar' into 'BAR'
                if 'Bar' in df.columns:
                    if 'BAR' not in df.columns:
                        df['BAR'] = df['Bar']
                    else:
                        df['BAR'] = df['BAR'].fillna(df['Bar'])
                        try:
                             mask = (df['BAR'] == "") & (df['Bar'] != "")
                             df.loc[mask, 'BAR'] = df.loc[mask, 'Bar']
                        except:
                             pass

                # Normalize BAR values (remove .0)
                if 'BAR' in df.columns:
                    df['BAR'] = df['BAR'].astype(str).str.replace(r'\.0$', '', regex=True).replace('nan', '')

                for col in self.columns:
                    if col not in df.columns:
                        df[col] = ""
                return df.fillna("")
            except:
                pass
        
        return pd.DataFrame(columns=self.columns)

    def integrate_master_db(self):
        """Scans MIDI_Library/*/*.xlsx and creates MasterLibraly.xlsx"""
        # scan
        if not os.path.exists(self.midi_lib_path):
            return
            
        dfs = []
        # Glob pattern: MIDI_Library/**/*.xlsx (Recursive)
        pattern = os.path.join(self.midi_lib_path, "**", "*.xlsx")
        files = glob.glob(pattern, recursive=True)
        
        if not files:
            return
            
        for f in files:
            # Skip conflict files or temps
            if "~" in f: continue
            
            try:
                d = pd.read_excel(f)
                
                # Consolidation: Merge 'Bar' into 'BAR'
                if 'Bar' in d.columns:
                    if 'BAR' not in d.columns:
                        d['BAR'] = d['Bar']
                    else:
                        d['BAR'] = d['BAR'].fillna(d['Bar'])
                        try:
                             mask = (d['BAR'] == "") & (d['Bar'] != "")
                             d.loc[mask, 'BAR'] = d.loc[mask, 'Bar']
                        except:
                             pass
                
                # Normalize BAR values (remove .0)
                if 'BAR' in d.columns:
                    d['BAR'] = d['BAR'].astype(str).str.replace(r'\.0$', '', regex=True).replace('nan', '')
                
                # Normalize
                for col in self.columns:
                    if col not in d.columns:
                        d[col] = ""
                # Add Source Info if needed, but maybe not strictly required for Master view if paths are absolute/correct relative
                # Ensure FilePath is usable.
                # If they are relative to Project Root, they are fine.
                dfs.append(d)
            except Exception as e:
                print(f"Error integrating {f}: {e}")
                
        if dfs:
            combined = pd.concat(dfs, ignore_index=True)
            # Dedup? Maybe by FilePath?
            combined.drop_duplicates(subset=["FilePath"], keep="last", inplace=True)
            
            master_dest = os.path.join(self.root_dir, "MasterLibraly.xlsx")
            try:
                combined.to_excel(master_dest, index=False)
                # Update self.df to reflect this fresh integration
                self.df = combined.fillna("")
                print("Master DB integrated successfully.")
            except PermissionError:
                print(f"Warning: Could not write to {master_dest}. File is open in Excel. Skipping integration save.")
                # We can still use the combined data in memory if we want, or fallback?
                # Usually better to use what we read, even if we can't save the consolidated version.
                self.df = combined.fillna("")
            except Exception as e:
                 print(f"Error saving Master DB: {e}")

    def save_db(self):
        # We only save NEW entries to the LOCAL DB.
        # But self.df contain EVERYTHING.
        # This is tricky without tracking "New/Modified" rows.
        # Simplified approach:
        # We don't save self.df back to Master. We leave Master as read-generated.
        # We ONLY append properly in `add_entry`.
        pass
        
    def _append_to_local(self, row_dict):
        # Helper to append just one row to local excel with Dropdowns
        import openpyxl
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.utils import get_column_letter
        import ui_constants as C
        
        try:
            # 1. Update Data (Pandas)
            if os.path.exists(self.local_db_path):
                try:
                    local_df = pd.read_excel(self.local_db_path)
                except:
                    local_df = pd.DataFrame(columns=self.columns)
            else:
                local_df = pd.DataFrame(columns=self.columns)
            
            new_df = pd.DataFrame([row_dict])
            # Align columns
            for col in self.columns:
                if col not in new_df.columns:
                    new_df[col] = ""
                    
            combined = pd.concat([local_df, new_df], ignore_index=True)
            combined.to_excel(self.local_db_path, index=False)
            
            # 2. Add Validations (OpenPyXL)
            wb = openpyxl.load_workbook(self.local_db_path)
            ws = wb.active
            
            # Helper to add validation
            def add_dv(col_name, options_list):
                 # Find Col Index
                try:
                    col_idx = combined.columns.get_loc(col_name) + 1 # 1-based
                    col_letter = get_column_letter(col_idx)
                except KeyError:
                    return # Column not found
                
                # Check for "Lists" sheet for long lists
                if len(",".join(options_list)) > 255:
                     list_sheet_name = "_Lists"
                     if list_sheet_name not in wb.sheetnames:
                         wb.create_sheet(list_sheet_name)
                     lws = wb[list_sheet_name]
                     
                     # Write options to column (e.g. A for this list)
                     # Simple hash to separate lists? 
                     # For now, let's just dump Instruments to A
                     if col_name == "Instruments":
                         for i, opt in enumerate(options_list):
                             lws.cell(row=i+1, column=1, value=opt)
                         formula = f"='{list_sheet_name}'!$A$1:$A${len(options_list)}"
                     else:
                         return # Handle other long lists if any
                     
                     dv = DataValidation(type="list", formula1=formula, allow_blank=True)
                else:
                    dv = DataValidation(type="list", formula1=f'"{",".join(options_list)}"', allow_blank=True)
                
                # Apply to entire column (or just used range)
                # Apply from row 2 to max_row + 100
                ws.add_data_validation(dv)
                dv.add(f"{col_letter}2:{col_letter}1000")

            # Apply Validations
            add_dv("Root", C.KEY_LIST) # Mapped from text "Key" in user request, but column is Root? User said "Key". 
                                       # In columns we have "Root". Assuming User "Key" -> DB "Root".
            add_dv("Category", C.CATEGORY_LIST)
            add_dv("Instruments", C.INSTRUMENT_LIST)
            add_dv("Chord", C.CHORD_LIST)
            
            wb.save(self.local_db_path)
            
        except Exception as e:
            print(f"Error saving to local db {self.local_db_path}: {e}")

    def add_entry(self, metadata):
        # Ensure new entry has all columns
        row = {col: metadata.get(col, "") for col in self.columns}
        
        # Assign Source = Local DB
        row["_SourceFile"] = self.local_db_name
        
        # Handle RelPath
        abs_path = metadata.get("FilePath")
        if abs_path and os.path.isabs(abs_path):
            try:
                rel_path = os.path.relpath(abs_path, self.root_dir)
                row["FilePath"] = rel_path
            except ValueError:
                row["FilePath"] = abs_path
        
        new_row_df = pd.DataFrame([row])
        self.df = pd.concat([self.df, new_row_df], ignore_index=True)
        # Ensure BAR is filled if Bar was somehow passed (though row construction filters it)
        # But row is constructed from self.columns, so Bar would be ignored if passed in metadata
        # unless we explicitly map it.
        # It's better to assume metadata has "BAR". User input handling should ensure "BAR".
        
        # Save to local persistence
        self._append_to_local(row)

    def get_filtered_data(self, filters):
        # Return a Filtered VIEW of the DF
        # We don't modify self.df here
        if self.df.empty:
            return self.df
            
        result = self.df.copy()
        
        for key, value in filters.items():
            if value and value != "No" and value != "None":
                # Assuming exact match for now, or use str.contains?
                # Using exact match for Radio buttons usually.
                # Ensure column exists
                if key in result.columns:
                    # Filter
                    result = result[result[key].astype(str) == value]
        
        return result
