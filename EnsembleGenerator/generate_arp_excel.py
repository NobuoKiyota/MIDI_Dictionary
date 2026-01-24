import pandas as pd
import os
from openpyxl.styles import PatternFill, Alignment, Font

def create_excel_with_formatting():
    file_path = "arpeggio_patterns.xlsx"
    
    # --- Sheet 1: Readme ---
    readme_data = [
        {"Item": "style_name", "Description": "スタイル名。Generatorで指定する際のIDになります。 (例: LoFi_01)"},
        {"Item": "swing", "Description": "スウィング感。偶数ステップ(裏拍)の開始時間を遅らせる秒数。 (例: 0.03 = 30ms)"},
        {"Item": "s01 - s16", "Description": "Sequencer Steps. ノートインデックス。\n・0, 1, 2... : コード構成音番号 (0=Root)\n・r : 休符 (Rest)\n・複数指定 (例: \"0,1\") : ストラミング"},
        {"Item": "g01 - g16", "Description": "Gate (音価)。整数指定。プログラム側で 1/10 されます。\n・10 : 1.0 (16分音符)\n・5 : 0.5 (スタッカート)\n・20 : 2.0 (Tie)"},
        {"Item": "v01 - v16", "Description": "Velocity Multiplier。整数指定。プログラム側で 1/10 されます。\n・10 : 1.0 (標準)\n・12 : 1.2 (アクセント)\n・8 : 0.8 (弱)"},
        {"Item": "Octave Wrapping", "Description": "コード音数超えでオクターブシフト。"}
    ]
    df_readme = pd.DataFrame(readme_data)
    
    # --- Sheet 2: Patterns ---
    data = []
    
    def to_int(lst):
        return [int(x * 10) for x in lst]

    # 1. LoFi_01
    seq = [0, 'r', 1, 2, 'r', 0, 1, 3, 0, 'r', 1, 2, 'r', 0, 1, 3]
    gate_flt = [1.0, 0, 1.2, 1.2, 0, 1.0, 1.2, 1.2, 1.0, 0, 1.2, 1.2, 0, 1.0, 1.2, 1.2]
    vel_flt = [0.8, 0, 0.9, 0.7, 0, 0.8, 0.9, 0.7, 0.8, 0, 0.9, 0.7, 0, 0.8, 0.9, 0.7]
    
    gate = to_int(gate_flt)
    vel = to_int(vel_flt)
    
    row_lofi = {"style_name": "LoFi_01", "swing": 0.03}
    for i in range(16):
        row_lofi[f"s{i+1:02}"] = seq[i]
        row_lofi[f"g{i+1:02}"] = gate[i]
        row_lofi[f"v{i+1:02}"] = vel[i]
    data.append(row_lofi)
    
    # 2. Trance_Gate
    seq = [0, 1, 2, 0, 1, 2, 0, 1, 2, 0, 1, 2, 0, 1, 2, 0]
    gate = [6] * 16
    vel = [12, 8, 10, 8] * 4
    
    row_trance = {"style_name": "Trance_Gate", "swing": 0.0}
    for i in range(16):
        row_trance[f"s{i+1:02}"] = seq[i]
        row_trance[f"g{i+1:02}"] = gate[i]
        row_trance[f"v{i+1:02}"] = vel[i]
    data.append(row_trance)
    
    # 3. Strum_Slow
    seq = ["0,1,2", 'r', 'r', 'r'] * 4
    gate = [20, 0, 0, 0] * 4
    vel = [10, 0, 0, 0] * 4
    
    row_strum = {"style_name": "Strum_Slow", "swing": 0.0}
    for i in range(16):
        row_strum[f"s{i+1:02}"] = seq[i]
        row_strum[f"g{i+1:02}"] = gate[i]
        row_strum[f"v{i+1:02}"] = vel[i]
    data.append(row_strum)
    
    # Columns
    cols = ["style_name", "swing"]
    cols += [f"s{i+1:02}" for i in range(16)]
    cols += [f"g{i+1:02}" for i in range(16)]
    cols += [f"v{i+1:02}" for i in range(16)]
    
    df_patterns = pd.DataFrame(data, columns=cols)
    
    try:
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df_patterns.to_excel(writer, sheet_name='Patterns', index=False)
            df_readme.to_excel(writer, sheet_name='Readme', index=False)
            
            # --- API Formatting ---
            ws = writer.sheets['Patterns']
            from openpyxl.utils import get_column_letter

            # Colors (Pastel)
            # Sequence (sXX): Light Blue (E0F7FA) -> openpyxl uses ARGB hex (FF + Hex)
            # Gate (gXX): Light Green (F1F8E9)
            # Velocity (vXX): Light Orange/Pink (FFF3E0)
            
            fill_seq = PatternFill(start_color="FFE0F7FA", end_color="FFE0F7FA", fill_type="solid")
            fill_gate = PatternFill(start_color="FFF1F8E9", end_color="FFF1F8E9", fill_type="solid")
            fill_vel = PatternFill(start_color="FFFFF3E0", end_color="FFFFF3E0", fill_type="solid")
            
            center_align = Alignment(horizontal='center', vertical='center')
            
            # Row count: 1 header + len(data) rows
            max_row = len(data) + 1
            max_col = len(cols) # 2 + 16*3 = 50
            
            # 1. Column Widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 8
            for c in range(3, 51): # C onwards
                ws.column_dimensions[get_column_letter(c)].width = 5.0
                
            # 2. Iterate all cells
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    
                    # Alignment: All centered except Col A (Style Name)
                    if col > 1:
                        cell.alignment = center_align
                        
                    # Colors
                    # C(3) - R(18): Sequence (s01-s16)
                    if 3 <= col <= 18:
                        cell.fill = fill_seq
                    # S(19) - Ah(34): Gate (g01-g16)
                    elif 19 <= col <= 34:
                        cell.fill = fill_gate
                    # Ai(35) - Ax(50): Velocity (v01-v16)
                    elif 35 <= col <= 50:
                        cell.fill = fill_vel
                        
        print(f"Successfully created {file_path} with Colors and Centering.")
    except Exception as e:
        print(f"Error creating Excel: {e}")

if __name__ == "__main__":
    create_excel_with_formatting()
